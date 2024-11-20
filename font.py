import  openpyxl  as  op
from  openpyxl.styles  import Font, Border, Side, Alignment, PatternFill, Protection
from  openpyxl.styles.colors  import  Color


#Workbook 및 Worksheet 객체 설정하기
wb = op.Workbook()
ws = wb.active


#font test1 : 직접 font 설정하기
ws["A1"].value = "Font test1"
ws["A1"].font = Font(size=20, italic = True, bold = True)

#font test2 : format을 정해놓고 font 설정하기
ws["A2"].value = "Font Test2"
font_format = Font(size=12, name='굴림', color = 'FF000000')
ws["A2"].font = font_format


#"C3"에 값 입력
ws["C3"].value = "1개 Cell"

#border test1 : 위에는 실선, 아래에는 이중선 적용 예시 코드
ws["C3"].border = Border(top = Side(border_style="thin", color="000000"), bottom = Side(border_style="double"))


# "C2"와 "C4"에 Text 입력  
ws["C2"].value = "Alignment test1"
ws["C4"].value = "Alignment test2"

# 셀 너비, 높이 설정하기
ws.row_dimensions[2].height = 50 #2행의 높이 50으로
ws.row_dimensions[4].height = 50 #4행의 높이 50으로
ws.column_dimensions['C'].width = 50 #C열의 너비 50으로

#Alignment test1 가로 & 세로 나열
ws["C2"].alignment = Alignment(horizontal = 'left', vertical='center')

#Alignment test2 가로 & 세로 나열
format1 = Alignment(horizontal = 'center', vertical='center')
ws["C4"].alignment = format1


#PatternFill test1 : green
ws["C1"].fill = PatternFill(fill_type='solid', fgColor="00FF00")

#PatternFill test2 : Black
#ws["C5"].fill = PatternFill(fill_type='solid', fgColor="000000")

#text 입력하기
ws["C3"].value = "Protection test1 : locked"
ws["C5"].value = "Protection test2 : hidden"

#Protection 속성 설정하기(숨김/잠금)
ws["C3"].protection = Protection(locked = True, hidden = True)
ws["C5"].protection = Protection(locked = False, hidden = False)


#Workbook(엑셀) 저장 및 객체 닫기
wb.save("font.xlsx")
wb.close()
