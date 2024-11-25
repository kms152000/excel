# openpyxl 모듈 import
import  openpyxl  as  op  

wb = op.load_workbook("test.xlsx")
ws = wb["sheet1"]
# ws = wb.active  #활성화 되어있는 시트 설정


# Cell data 읽기
data1 = ws.cell(row=1, column=2).value
print(data1)

data2 = ws["A1:C3"]
for data in data2:
	for cell_data in data:
		print(cell_data.value, end=" ") 
print()

# Cell data 쓰기
ws = wb["sheet2"]
ws.cell(row = 1, column = 1).value = "가"
ws["B1"].value = "나"
ws["C1"].value = "다"

ws = wb["sheet"]
datalist = [2,4,8,16,32,64,128,256]
i = 1
for data in datalist:
	ws.cell(row=i,column=1).value = data
	i += 1


# Cell data 삭제하기
rng = ws["A1:H8"] # 구간 정하기
for row_data in rng:
	for cell in row_data:
		if cell.column > 1:
			cell.value = None # ""으로 정의할 경우 데이터가 있는 것으로 간주
ws.delete_rows(9,2) # 9행부터 2개        
ws.delete_cols(9,2)  
# wb.remove(ws) # 해당 시트 삭제
# wb.create_sheet("sheet") # 해당 시트 재생성


# rows, columns 속성 이해하기
#ws = wb.create_sheet["sheet3"]
ws = wb["sheet3"]
data_list = [1, 2, 3, 4, 5, 6, 7, 8, 9]
i = 1 # 행
j = 1 # 열
for data in data_list:
	ws.cell(row = i, column = j, value = data)
	j += 1
	if j > 3:
		i += 1
		j = 1

print(ws["A1:C3"]) # 튜플

print("#rows 출력")
for  row_rng  in  ws.rows:
    print(row_rng) # 1차원 배열
print("#columns 출력")
for  col_rng  in  ws.columns:
    print(col_rng)



wb.save("test.xlsx")