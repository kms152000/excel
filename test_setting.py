import openpyxl as op

# wb = op.Workbook() # 새로운 Workbook 객체 생성
# wb.save("test.xlsx")
# wb.save("D:/MyProj/Python/excel/test.xlsx") # 엑셀 파일 생성

path = r"D:\MyProj\Python\excel"
wb = op.load_workbook(path + "/test.xlsx")

# ws1 = wb.create_sheet("sheet1")  # 기본적으로 시트는 마지막에 추가됨
# ws2 = wb.create_sheet("sheet2")
# ws3 = wb.create_sheet("sheet3")


wb.save("test.xlsx")
wb.save("D:/MyProj/Python/excel/test.xlsx")

ws_list = wb.sheetnames  #해당 Workbook의 시트 목록을 리스트로 저장
for  sht  in  ws_list:
    ws = wb[sht] #Sheet 객체 생성
    print(ws) #객체 출력